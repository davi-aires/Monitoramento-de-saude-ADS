"""
Arquivo auxiliar para gerenciar o banco de dados
Use este arquivo para operações de manutenção
"""

import sqlite3
import os
from dotenv import load_dotenv

load_dotenv()
from datetime import datetime
from zoneinfo import ZoneInfo
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, "database.db")
DATABASE_URL = os.getenv('DATABASE_URL')  # Render define automaticamente

BR_TZ = ZoneInfo('America/Sao_Paulo')

def agora_br():
    """Datetime atual no fuso de Brasília, sem tzinfo"""
    return datetime.now(BR_TZ).replace(tzinfo=None)

def ph():
    """Placeholder para query: '?' no SQLite, '%s' no PostgreSQL"""
    return '%s' if DATABASE_URL else '?'

def get_db_connection():
    """Conexão com SQLite local ou PostgreSQL (Render)"""
    if DATABASE_URL:
        import psycopg2
        import psycopg2.extras
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def get_cursor(conn):
    """Cursor com acesso por nome de coluna para SQLite e PostgreSQL"""
    if DATABASE_URL:
        import psycopg2.extras
        return conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    return conn.cursor()

def init_db():
    """Inicializa o banco de dados (SQLite ou PostgreSQL)"""
    if DATABASE_URL:
        import psycopg2
        conn = psycopg2.connect(DATABASE_URL)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS saude (
                id SERIAL PRIMARY KEY,
                data_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                nome TEXT NOT NULL,
                consumo_agua INTEGER NOT NULL,
                minutos_sol INTEGER NOT NULL,
                pratica_exercicio TEXT NOT NULL,
                humor TEXT NOT NULL
            )
        ''')
        conn.commit()
        cursor.close()
        conn.close()
        print("✅ PostgreSQL: tabela verificada/criada!")
    elif not os.path.exists(DATABASE):
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS saude (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                data_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                nome TEXT NOT NULL,
                consumo_agua INTEGER NOT NULL,
                minutos_sol INTEGER NOT NULL,
                pratica_exercicio TEXT NOT NULL,
                humor TEXT NOT NULL
            )
        ''')
        conn.commit()
        conn.close()
        print("✅ Banco de dados SQLite criado com sucesso!")
    else:
        print("ℹ️  Banco de dados já existe!")

def ver_todos_registros():
    """Exibe todos os registros do banco"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM saude ORDER BY data_registro DESC')
    registros = cursor.fetchall()
    conn.close()

    if registros:
        print("\n📊 REGISTROS SALVOS:\n")
        print(f"{'ID':<4} {'Data':<20} {'Água':<8} {'Sol':<8} {'Exercício':<12} {'Humor':<10}")
        print("-" * 70)
        for registro in registros:
            print(f"{registro['id']:<4} {registro['nome']:<20} {registro['data_registro']:<20} {registro['consumo_agua']:<8} {registro['minutos_sol']:<8} {registro['pratica_exercicio']:<12} {registro['humor']:<10}")
        print(f"\nTotal de registros: {len(registros)}")
    else:
        print("⚠️  Nenhum registro encontrado!")

def limpar_banco():
    """Limpa todos os registros do banco (use com cuidado!)"""
    resposta = input("\n⚠️  Tem certeza que deseja APAGAR todos os registros? (S/N): ").upper()

    if resposta == 'S':
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM saude')
        conn.commit()
        conn.close()
        print("✅ Todos os registros foram apagados!")
    else:
        print("❌ Operação cancelada!")

def resetar_banco():
    """Apaga o banco de dados completamente"""
    resposta = input("\n⚠️  Tem certeza que deseja RESETAR o banco de dados? (S/N): ").upper()

    if resposta == 'S':
        if os.path.exists(DATABASE):
            os.remove(DATABASE)
            init_db()
            print("✅ Banco de dados resetado com sucesso!")
        else:
            print("ℹ️  Banco de dados não existe!")
    else:
        print("❌ Operação cancelada!")

def limpar_registro(id):
    """Limpa apenas um registro"""
    resposta = input(f"\n⚠️  Tem certeza que deseja APAGAR o registro de ID = {id}: ").upper()

    if resposta == 'S':
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM saude where id = ?', (id,))
        conn.commit()
        conn.close()
        print("✅ Registro apagado com sucesso!")
    else:
        print("❌ Operação cancelada!")

def gerar_relatorio_excel(caminho_arquivo=None):
    """Gera um relatório Excel com todos os registros do banco de dados"""
    conn = get_db_connection()
    cursor = get_cursor(conn)
    cursor.execute('SELECT * FROM saude ORDER BY data_registro DESC')
    registros = cursor.fetchall()
    conn.close()

    if not registros:
        print("⚠️  Nenhum registro encontrado para gerar relatório!")
        return None

    if caminho_arquivo is None:
        timestamp = agora_br().strftime("%Y%m%d_%H%M%S")
        caminho_arquivo = os.path.join(BASE_DIR, f"relatorio_saude_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monitoramento de Saúde"

    verde_escuro = "1F6B3B"
    cinza_claro  = "F2F2F2"

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill  = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # Título
    ws.merge_cells("A1:G1")
    ws["A1"].value     = "🏥 Relatório de Monitoramento de Saúde"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=16, color=verde_escuro)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"].value     = f"Gerado em: {agora_br().strftime('%d/%m/%Y às %H:%M:%S')}"
    ws["A2"].font      = Font(name="Calibri", italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    # Cabeçalhos
    colunas = ["ID", "Data de Registro", "Nome", "Água (ml)", "Sol (min)", "Exercício", "Humor"]
    for col_idx, col_nome in enumerate(colunas, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_nome)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 22

    # Dados
    for row_idx, registro in enumerate(registros, start=5):
        fill_color = cinza_claro if row_idx % 2 == 0 else "FFFFFF"
        row_fill   = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        valores = [registro["id"], registro["data_registro"], registro["nome"],
                   registro["consumo_agua"], registro["minutos_sol"],
                   registro["pratica_exercicio"], registro["humor"]]
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill      = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border
            cell.font      = Font(name="Calibri", size=11)

    # Total
    total_row = len(registros) + 5
    ws.merge_cells(f"A{total_row}:B{total_row}")
    ws[f"A{total_row}"].value     = f"Total de registros: {len(registros)}"
    ws[f"A{total_row}"].font      = Font(name="Calibri", bold=True, size=11, color=verde_escuro)
    ws[f"A{total_row}"].alignment = Alignment(horizontal="left", vertical="center")

    # Largura das colunas
    for col_idx, largura in enumerate([6, 22, 20, 12, 12, 14, 12], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    wb.save(caminho_arquivo)
    print(f"✅ Relatório Excel gerado: {caminho_arquivo}")
    return caminho_arquivo


def importar_excel(caminho_arquivo):
    """Importa registros de um Excel gerado pelo sistema para o banco de dados.
    Ignora registros cujo ID já exista no banco."""
    if not os.path.exists(caminho_arquivo):
        print(f"❌ Arquivo não encontrado: {caminho_arquivo}")
        return

    init_db()  # garante que a tabela existe antes de importar

    wb = openpyxl.load_workbook(caminho_arquivo)
    ws = wb.active

    conn = get_db_connection()
    cursor = get_cursor(conn)

    importados = 0
    ignorados = 0

    # Dados começam na linha 5 (1=título, 2=data geração, 3=vazia, 4=cabeçalho)
    for row in ws.iter_rows(min_row=5, values_only=True):
        id_val, data_registro, nome, consumo_agua, minutos_sol, pratica_exercicio, humor = row

        # Linha do total (última linha) — pula
        if id_val is None or str(id_val).startswith("Total"):
            continue

        # Converte data se vier como string
        if isinstance(data_registro, str):
            try:
                data_registro = datetime.strptime(data_registro, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    data_registro = datetime.strptime(data_registro, "%d/%m/%Y %H:%M:%S")
                except ValueError:
                    data_registro = agora_br()

        try:
            cursor.execute(
                f'INSERT INTO saude (data_registro, nome, consumo_agua, minutos_sol, pratica_exercicio, humor) '
                f'VALUES ({ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()})',
                (data_registro, str(nome), int(consumo_agua), int(minutos_sol),
                 str(pratica_exercicio), str(humor))
            )
            importados += 1
        except Exception as e:
            print(f"⚠️  Linha ignorada (erro: {e})")
            ignorados += 1

    conn.commit()
    cursor.close()
    conn.close()
    print(f"✅ Importação concluída: {importados} registros inseridos, {ignorados} ignorados.")




def migrar_sqlite_para_postgres():
    """Exporta todos os dados do SQLite local para o PostgreSQL do Render.
    Configure DATABASE_URL no .env antes de rodar."""
    if not DATABASE_URL:
        print("❌ Defina DATABASE_URL no .env com a URL do PostgreSQL do Render.")
        return
    if not os.path.exists(DATABASE):
        print("❌ Banco SQLite não encontrado.")
        return

    sqlite_conn = sqlite3.connect(DATABASE)
    sqlite_conn.row_factory = sqlite3.Row
    cur_sqlite = sqlite_conn.cursor()
    cur_sqlite.execute('SELECT * FROM saude ORDER BY id')
    registros = cur_sqlite.fetchall()
    sqlite_conn.close()

    if not registros:
        print("⚠️  Nenhum registro no SQLite para migrar.")
        return

    import psycopg2
    pg_conn = psycopg2.connect(DATABASE_URL)
    pg_cur = pg_conn.cursor()
    pg_cur.execute('''
        CREATE TABLE IF NOT EXISTS saude (
            id SERIAL PRIMARY KEY,
            data_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            nome TEXT NOT NULL,
            consumo_agua INTEGER NOT NULL,
            minutos_sol INTEGER NOT NULL,
            pratica_exercicio TEXT NOT NULL,
            humor TEXT NOT NULL
        )
    ''')
    migrados = 0
    for r in registros:
        pg_cur.execute(
            'INSERT INTO saude (data_registro, nome, consumo_agua, minutos_sol, pratica_exercicio, humor) '
            'VALUES (%s, %s, %s, %s, %s, %s)',
            (r['data_registro'], r['nome'], r['consumo_agua'],
             r['minutos_sol'], r['pratica_exercicio'], r['humor'])
        )
        migrados += 1
    pg_conn.commit()
    pg_cur.close()
    pg_conn.close()
    print(f"✅ {migrados} registros migrados com sucesso para o PostgreSQL!")



def menu():
    """Menu principal"""
    while True:
        print("\n" + "="*50)
        print("🏥 GERENCIADOR DE BANCO DE DADOS - SAÚDE")
        print("="*50)
        print("\n1. Ver todos os registros")
        print("2. Limpar apenas um registro")
        print("3. Limpar registros (manter banco)")
        print("4. Resetar banco de dados")
        print("5. Gerar relatório Excel")
        print("6. Importar registros de Excel")
        print("7. Migrar SQLite → PostgreSQL (Render)")
        print("8. Sair")

        opcao = input("\nEscolha uma opção (1-8): ")

        if opcao == '1':
            ver_todos_registros()
        elif opcao == '2':
            id = input("\nDigite o ID do registro que deseja apagar: ")
            limpar_registro(id)
        elif opcao == '3':
            limpar_banco()
        elif opcao == '4':
            resetar_banco()
        elif opcao == '5':
            gerar_relatorio_excel()
        elif opcao == '6':
            arquivo = input("\nCaminho do arquivo Excel: ").strip().strip('"')
            importar_excel(arquivo)
        elif opcao == '7':
            migrar_sqlite_para_postgres()
        elif opcao == '8':
            print("\n👋 Até logo!")
            break
        else:
            print("❌ Opção inválida!")

if __name__ == '__main__':
    menu()


